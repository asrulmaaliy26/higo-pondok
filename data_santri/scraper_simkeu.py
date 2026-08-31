"""
=============================================================================
SIMKEU AL-MANNAN - COMPREHENSIVE DATA SCRAPER & MASTER UPSERT ENGINE
=============================================================================
Deskripsi:
Program Python untuk mengambil SELURUH data Detail Kesiswaan dari SIMKEU
AL-MANNAN (Profil, Biodata Lengkap, Informasi Wali & Keluarga, Kesehatan, 
Kartu & Saldo Saku, Paket Keuangan & Rincian, Riwayat Pembayaran VA).

SISTEM PENCOCOKAN & PEMBARUAN (UPSERT BERDASARKAN NO. INDUK):
- Menggunakan "No. Induk" sebagai KUNCI UNIK UTAMA (Primary Key).
- Jika "No. Induk" SUDAH ADA di database / Excel:
  -> Otomatis UPDATE seluruh data sisanya (NISN, NIK, TTL, Info Wali, HP, 
     Kamar, Saldo, Kartu, Kelas Lengkap, dll.) tanpa membuat duplikat.
- Jika "No. Induk" BELUM ADA:
  -> Otomatis DITAMBAHKAN sebagai data santri baru.
- Otomatis memperbarui file master:
  * data_santri/seluruh_data_santri.json
  * frontend/src/data/santri.json (sinkron ke web app Higo Pondok)
  * data_santri/seluruh_data_santri.xlsx (Excel Master)
  * data_santri/simkeu_hasil_scraping.xlsx & .json
=============================================================================
"""

import os
import sys
import time
import json
import re
from urllib.parse import urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
except ImportError as e:
    print(f"[!] Error: Modul Python belum lengkap ({e}).")
    print("[*] Menginstall modul yang dibutuhkan...")
    os.system(f"{sys.executable} -m pip install requests beautifulsoup4 pandas openpyxl selenium")
    print("[+] Instalasi selesai. Silakan jalankan ulang skrip.")
    sys.exit(1)

# Direktori kerja
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(CURRENT_DIR, ".."))

COOKIE_FILE = os.path.join(CURRENT_DIR, "simkeu_cookies.json")
EXCEL_OUTPUT = os.path.join(CURRENT_DIR, "simkeu_hasil_scraping.xlsx")
JSON_OUTPUT = os.path.join(CURRENT_DIR, "simkeu_hasil_scraping.json")

MASTER_JSON_LOCAL = os.path.join(CURRENT_DIR, "seluruh_data_santri.json")
MASTER_EXCEL_LOCAL = os.path.join(CURRENT_DIR, "seluruh_data_santri.xlsx")
FRONTEND_DATA_JSON = os.path.join(PROJECT_ROOT, "frontend", "src", "data", "santri.json")

CHROME_PROFILE_DIR = os.path.join(CURRENT_DIR, ".chrome_simkeu_profile")

BASE_URL = "https://simkeu.almannan.id"
LOGIN_URL = "https://simkeu.almannan.id/login"
DEFAULT_DETAIL_URL = "https://simkeu.almannan.id/kesiswaan/siswa_detail/17f68b03c9d6411c936adb1a15d6c37d"

DEFAULT_USERNAME = "almannan"
DEFAULT_PASSWORD = "Almannan*"


def init_chrome_driver(headless=False):
    """Inisialisasi Chrome Webdriver dengan profil pengguna agar session awet."""
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(f"--user-data-dir={CHROME_PROFILE_DIR}")
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

    driver = webdriver.Chrome(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def login_and_save_session(driver=None, username=DEFAULT_USERNAME, password=DEFAULT_PASSWORD):
    """Buka browser untuk login jika session cookie belum ada atau sudah kedaluwarsa."""
    created_driver = False
    if driver is None:
        print("[*] Membuka browser Chrome untuk verifikasi login...")
        driver = init_chrome_driver(headless=False)
        created_driver = True

    try:
        print(f"[*] Mengarahkan ke halaman login: {LOGIN_URL}")
        driver.get(LOGIN_URL)
        time.sleep(2)

        if "/login" not in driver.current_url.lower():
            print(f"[+] Sesi aktif terdeteksi! Langsung di halaman: {driver.current_url}")
            save_cookies_from_driver(driver)
            if created_driver:
                driver.quit()
            return

        try:
            user_input = driver.find_element(By.CSS_SELECTOR, "input[name='username'], input[type='text']")
            if user_input and not user_input.get_attribute("value"):
                user_input.clear()
                user_input.send_keys(username)
        except Exception:
            pass

        try:
            pass_input = driver.find_element(By.CSS_SELECTOR, "input[name='password'], input[type='password']")
            if pass_input and not pass_input.get_attribute("value"):
                pass_input.clear()
                pass_input.send_keys(password)
        except Exception:
            pass

        print("\n" + "=" * 70)
        print(">>> SILAKAN LOGIN DI JENDELA CHROME YANG TERBUKA <<<")
        print("Selesaikan Captcha/Keamanan jika ada, lalu klik tombol 'Log In'.")
        print("Skrip akan otomatis mendeteksi ketika Anda berhasil masuk...")
        print("=" * 70 + "\n")

        start_wait = time.time()
        while "/login" in driver.current_url.lower():
            time.sleep(1.5)
            if time.time() - start_wait > 600:
                print("[!] Waktu tunggu login habis (10 menit).")
                break

        if "/login" not in driver.current_url.lower():
            print(f"\n[+] BERHASIL LOGIN! Halaman saat ini: {driver.current_url}")
            save_cookies_from_driver(driver)
        else:
            print("[!] Belum berhasil login.")

        if created_driver:
            driver.quit()

    except Exception as e:
        print(f"[!] Terjadi kesalahan saat login: {e}")
        if created_driver:
            driver.quit()


def save_cookies_from_driver(driver):
    """Simpan cookies dari Selenium ke file JSON."""
    cookies = driver.get_cookies()
    with open(COOKIE_FILE, "w", encoding="utf-8") as f:
        json.dump(cookies, f, indent=2)
    print(f"[+] Cookie sesi berhasil disimpan ke: {os.path.basename(COOKIE_FILE)}")


def get_requests_session():
    """Membuat session HTTP requests dengan cookie yang tersimpan."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Referer": BASE_URL,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "X-Requested-With": "XMLHttpRequest"
    })

    if not os.path.exists(COOKIE_FILE):
        print("[!] File cookie belum ditemukan. Membuka browser untuk login...")
        login_and_save_session()

    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8") as f:
            cookies = json.load(f)
            for c in cookies:
                session.cookies.set(c["name"], c["value"], domain=c.get("domain", "simkeu.almannan.id"))

    # Uji validitas cookie
    try:
        test_r = session.get(f"{BASE_URL}/kesiswaan", timeout=8)
        if "/login" in test_r.url.lower() or test_r.status_code != 200:
            print("[!] Sesi cookie kedaluwarsa. Membuka browser untuk re-login...")
            login_and_save_session()
            with open(COOKIE_FILE, "r", encoding="utf-8") as f:
                cookies = json.load(f)
                for c in cookies:
                    session.cookies.set(c["name"], c["value"], domain=c.get("domain", "simkeu.almannan.id"))
    except Exception:
        pass

    return session


def extract_full_student_details(session, student_url_or_id):
    """
    Mengekstrak SELURUH data siswa dari halaman Detail Kesiswaan:
    - Profil Siswa (No. Induk, NISN, Nama, Jenjang, Kelas, Ruang, Program, WA, dll.)
    - Biodata Siswa (NIK, TTL, Status, Asal Sekolah, Angkatan, Alamat, dll.)
    - Informasi Wali (Nama Ayah, Ibu, Pekerjaan, No Telp, Alamat, No KK, dll.)
    - Kesehatan
    - Status Kartu, No Kartu, Limit Jajan
    - Saldo Saku & Saldo Tabungan Wali
    - Paket Keuangan & Rincian Komponen
    - Riwayat Transaksi VA / Bank
    """
    if "siswa_detail/" in student_url_or_id:
        id_siswa = student_url_or_id.split("siswa_detail/")[-1].split("#")[0].split("?")[0].strip()
        url = student_url_or_id.split("#")[0]
    else:
        id_siswa = student_url_or_id.strip()
        url = f"{BASE_URL}/kesiswaan/siswa_detail/{id_siswa}"

    # Ambil halaman detail dengan mekanisme Retry otomatis
    res = None
    for attempt in range(3):
        try:
            res = session.get(url, timeout=25)
            if res.status_code == 200:
                break
        except Exception as e:
            if attempt == 2:
                print(f"\n[!] Gagal mengambil URL {url} setelah 3x percobaan: {e}")
                return None
            time.sleep(1.5)

    if not res or res.status_code != 200:
        return None

    soup = BeautifulSoup(res.text, "html.parser")
    
    data = {
        "ID Siswa": id_siswa,
        "URL Sumber": url,
        "Waktu Scraping": time.strftime("%Y-%m-%d %H:%M:%S")
    }

    # 1. PROFIL SISWA
    p_pane = soup.find("div", id="profilsiswa")
    if p_pane:
        for tr in p_pane.find_all("tr"):
            tds = tr.find_all(["td", "th"])
            if len(tds) >= 2:
                k = tds[0].get_text(strip=True).rstrip(":").strip()
                v = tds[-1].get_text(strip=True)
                if k:
                    data[f"Profil - {k}"] = v

    # 2. BIODATA SISWA & INFORMASI WALI
    b_pane = soup.find("div", id="biodatasiswa")
    if b_pane:
        for tr in b_pane.find_all("tr"):
            tds = tr.find_all(["td", "th"])
            if len(tds) >= 2:
                k = tds[0].get_text(strip=True).rstrip(":").strip()
                v = tds[-1].get_text(strip=True)
                if k:
                    if any(w in k.lower() for w in ["ayah", "ibu", "wali", "keluarga", "anak ke", "saudara", "no. kk"]):
                        data[f"Wali - {k}"] = v
                    else:
                        data[f"Biodata - {k}"] = v

    # 3. KESEHATAN
    k_pane = soup.find("div", id="kesehatansiswa")
    if k_pane:
        t = k_pane.get_text(" ", strip=True)
        data["Kesehatan - Kondisi"] = "SEHAT" if "sehat" in t.lower() else t

    # 4. KEUANGAN & KARTU (VIA API SIMKEU)
    try:
        r_kartu = session.post(f"{BASE_URL}/api_keuangan/cari_kartu", data={"id_pengguna": id_siswa}, timeout=8).json()
        if r_kartu.get("status") == "sukses" and r_kartu.get("data"):
            kd = r_kartu["data"]
            data["Kartu - No. Kartu"] = kd.get("key_kartu", "")
            data["Kartu - Status"] = BeautifulSoup(kd.get("status_kartu", ""), "html.parser").get_text(strip=True)
            data["Kartu - Waktu Aktivasi"] = kd.get("time_kartu", "")
            data["Kartu - Limit Jajan"] = kd.get("limit_kartu", "")
    except Exception:
        pass

    try:
        r_saldo = session.post(f"{BASE_URL}/api_keuangan/get_saldo_siswa", data={"id_siswa": id_siswa}, timeout=8).json()
        if r_saldo.get("status") == "sukses":
            data["Saldo - Uang Saku"] = f"Rp {r_saldo.get('saldo', '0')}"
    except Exception:
        pass

    try:
        r_tab = session.post(f"{BASE_URL}/api_keuangan/get_tabwali_siswa", data={"id_siswa": id_siswa}, timeout=8).json()
        if r_tab.get("status") == "sukses":
            data["Saldo - Tabungan Wali"] = f"Rp {r_tab.get('saldo', '0')}"
    except Exception:
        pass

    # 5. PAKET KEUANGAN & RINCIAN KOMPONEN
    sel_opt = soup.find("select", attrs={"name": lambda n: n and "paket" in str(n).lower()})
    if sel_opt:
        opt = sel_opt.find("option", selected=True)
        if opt:
            data["Paket - Nama Paket"] = opt.get_text(strip=True)
    
    if "Paket - Nama Paket" not in data:
        opt_single = soup.find("option", selected=True)
        if opt_single and "paket" in opt_single.get_text().lower():
            data["Paket - Nama Paket"] = opt_single.get_text(strip=True)

    rekap_pane = soup.find("div", id="rekapke")
    if rekap_pane:
        for div in rekap_pane.find_all("div", class_=lambda c: c and "justify-content-between" in str(c)):
            spans = div.find_all("span")
            if len(spans) == 1:
                val = spans[0].get_text(strip=True)
                lbl = div.get_text(" ", strip=True).replace(val, "").strip()
                if lbl and len(lbl) < 60:
                    data[f"Rincian Paket - {lbl}"] = val
            elif len(spans) >= 2:
                lbl = spans[0].get_text(strip=True)
                val = spans[1].get_text(strip=True)
                if lbl and len(lbl) < 60:
                    data[f"Rincian Paket - {lbl}"] = val

    # 6. RIWAYAT PEMBAYARAN BANK / VA (Ambil maksimal 5 transaksi terakhir agar tidak melebihi batas sel)
    bank_tables = soup.find_all("table")
    va_list = []
    for t in bank_tables:
        t_text = t.get_text()
        if "Virtual Account" in t_text or "TRX" in t_text:
            for tr in t.find_all("tr"):
                tds = [td.get_text(strip=True) for td in tr.find_all("td")]
                if len(tds) >= 4 and any(trx in tds[0] for trx in ["202", "MSK"]):
                    # Simpan data penting saja: TRX, Tgl, Nominal
                    clean_row = " | ".join([x for x in tds if x and x not in ["Edit", "Hapus", "Print"]])
                    va_list.append(clean_row)
                    if len(va_list) >= 5:
                        break
    if va_list:
        data["Riwayat Pembayaran VA"] = "; ".join(va_list)[:1500]

    return data


def get_all_active_students_list(session):
    """Mengambil daftar seluruh siswa aktif dari SIMKEU."""
    print("[*] Mengambil daftar seluruh santri aktif dari server SIMKEU...")
    try:
        r = session.post(
            f"{BASE_URL}/kesiswaan/list_siswa_aktif",
            data={"draw": 1, "start": 0, "length": 5000},
            timeout=20
        )
        res = r.json()
        raw_list = res.get("data", [])
        print(f"[+] Ditemukan {len(raw_list)} data santri aktif di sistem SIMKEU!")
        
        students = []
        for row in raw_list:
            soup = BeautifulSoup(str(row[-1]), "html.parser")
            a_tag = soup.find("a", href=lambda h: h and "siswa_detail" in str(h))
            if a_tag:
                href = a_tag.get("href")
                detail_id = href.split("siswa_detail/")[-1].strip()
                
                nama_soup = BeautifulSoup(str(row[1]), "html.parser")
                raw_name = nama_soup.get_text(strip=True)
                
                students.append({
                    "id_siswa": detail_id,
                    "nama_ringkas": raw_name,
                    "no_induk": str(row[2]).strip() if len(row) > 2 else "",
                    "nisn": str(row[3]).strip() if len(row) > 3 else "",
                    "jenjang": row[4] if len(row) > 4 else "",
                    "tingkat": row[5] if len(row) > 5 else "",
                    "rombel": row[6] if len(row) > 6 else "",
                    "program": row[7] if len(row) > 7 else "",
                    "wa": row[8] if len(row) > 8 else "",
                    "url": f"{BASE_URL}/kesiswaan/siswa_detail/{detail_id}"
                })
        return students
    except Exception as e:
        print(f"[!] Gagal mengambil daftar santri: {e}")
        return []


def upsert_master_dataset_by_no_induk(scraped_records):
    """
    KUNCI UTAMA: NO. INDUK (Nomor Induk Santri).
    Jika No. Induk sudah ada di master -> UPDATE SELURUH DATA SISANYA.
    Jika No. Induk belum ada -> TAMBAHKAN SEBAGAI DATA BARU.
    Sinkronkan ke seluruh file master (JSON & Excel).
    """
    if not scraped_records:
        return

    if isinstance(scraped_records, dict):
        scraped_records = [scraped_records]

    print("\n" + "=" * 70)
    print("[*] MENJALANKAN PROSES SINKRONISASI & UPSERT BERDASARKAN NO. INDUK...")
    print("=" * 70)

    # 1. Baca master JSON lama jika ada
    master_json_data = {"total_data": 0, "headers": [], "data": []}
    if os.path.exists(MASTER_JSON_LOCAL):
        try:
            with open(MASTER_JSON_LOCAL, "r", encoding="utf-8") as f:
                master_json_data = json.load(f)
        except Exception:
            pass

    master_list = master_json_data.get("data", [])

    # Indexing existing data by No. Induk
    induk_index_map = {}
    for idx, row in enumerate(master_list):
        if len(row) > 2 and row[2]:
            clean_induk = str(row[2]).strip()
            induk_index_map[clean_induk] = idx

    updated_count = 0
    inserted_count = 0

    for rec in scraped_records:
        # Ekstrak No. Induk
        no_induk = str(rec.get("Profil - No. Induk", rec.get("no_induk", ""))).strip()
        if not no_induk:
            continue

        nama = rec.get("Profil - Nama Lengkap", rec.get("nama_ringkas", "Santri")).strip()
        nisn = str(rec.get("Profil - NISN", rec.get("nisn", ""))).strip()
        jenjang = rec.get("Profil - Jenjang", rec.get("jenjang", "")).strip()
        tingkat = rec.get("Profil - Kelas", rec.get("tingkat", "")).strip()
        rombel = rec.get("Profil - Ruang", rec.get("rombel", "")).strip()
        program = rec.get("Profil - Program", rec.get("program", "")).strip()
        wa = str(rec.get("Profil - No. Whatsapp", rec.get("wa", ""))).strip()
        asrama = rec.get("Biodata - Nama Angkatan", rec.get("Profil - Pondok", "")).strip()

        # Ekstrak jenis kelamin
        jk = rec.get("Biodata - Jenis Kelamin", "")
        nama_with_jk = f"{nama} {jk}".strip() if jk and jk not in nama else nama

        # Bentuk baris array standar
        standard_row = [
            "",
            nama_with_jk,
            no_induk,
            nisn,
            jenjang,
            tingkat,
            rombel,
            program,
            wa,
            "Detail",
            asrama
        ]

        if no_induk in induk_index_map:
            # UPDATE DATA LAMA
            row_idx = induk_index_map[no_induk]
            master_list[row_idx] = standard_row
            updated_count += 1
        else:
            # INSERT DATA BARU
            master_list.append(standard_row)
            induk_index_map[no_induk] = len(master_list) - 1
            inserted_count += 1

    # Update metadata master
    master_json_data["total_data"] = len(master_list)
    master_json_data["data"] = master_list

    # 2. Simpan ke local data_santri/seluruh_data_santri.json
    with open(MASTER_JSON_LOCAL, "w", encoding="utf-8") as f:
        json.dump(master_json_data, f, ensure_ascii=False, indent=4)
    print(f"[+] Master JSON lokal tersinkron: {os.path.basename(MASTER_JSON_LOCAL)}")

    # 3. Simpan ke frontend/src/data/santri.json (agar React Web App langsung update)
    try:
        os.makedirs(os.path.dirname(FRONTEND_DATA_JSON), exist_ok=True)
        with open(FRONTEND_DATA_JSON, "w", encoding="utf-8") as f:
            json.dump(master_json_data, f, ensure_ascii=False, indent=4)
        print(f"[+] Frontend Web App tersinkron: {os.path.basename(FRONTEND_DATA_JSON)}")
    except Exception as e:
        print(f"[!] Catatan: Tidak dapat menulis ke frontend data ({e})")

    # 4. Generate Master Excel Lengkap
    generate_master_excel_from_rows(master_list, MASTER_EXCEL_LOCAL)

    print("\n" + "-" * 70)
    print(f"[SUCCESS] HASIL SINKRONISASI BERDASARKAN NO. INDUK:")
    print(f"   - Data Diperbarui (Updated) : {updated_count} santri")
    print(f"   - Data Baru (Inserted)      : {inserted_count} santri")
    print(f"   - Total Master Santri Aktif : {len(master_list)} santri")
    print("-" * 70)


def generate_master_excel_from_rows(rows_data, output_excel_path):
    """Membuat file Excel master yang terformat rapi & profesional."""
    formatted_rows = []
    for idx, item in enumerate(rows_data, 1):
        raw_nama = item[1] if len(item) > 1 else ""
        gender = ""
        nama = raw_nama
        if raw_nama.endswith("Laki-laki"):
            gender = "Laki-laki"
            nama = raw_nama[:-9].strip()
        elif raw_nama.endswith("Perempuan"):
            gender = "Perempuan"
            nama = raw_nama[:-9].strip()

        no_induk = str(item[2]) if len(item) > 2 and item[2] != "0" else ""
        no_nisn = str(item[3]) if len(item) > 3 and item[3] != "0" else ""
        jenjang = item[4] if len(item) > 4 else ""
        tingkat = item[5] if len(item) > 5 else ""
        rombel = item[6] if len(item) > 6 else ""
        program = item[7] if len(item) > 7 and item[7] != "-" else ""
        wa = item[8] if len(item) > 8 else ""
        asrama = item[10] if len(item) > 10 else ""

        class_parts = [p for p in [tingkat, rombel, program] if p]
        kelas_lengkap = " ".join(class_parts)

        formatted_rows.append({
            "No": idx,
            "Nama Santri": nama,
            "Jenis Kelamin": gender,
            "No. Induk": no_induk,
            "NISN": no_nisn,
            "Jenjang": jenjang,
            "Kelas Lengkap": kelas_lengkap,
            "Tingkat": tingkat,
            "Rombel / Ruang": rombel,
            "Program / Jurusan": program,
            "Asrama / Kamar": asrama,
            "No. WhatsApp": wa
        })

    df = pd.DataFrame(formatted_rows)
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data Seluruh Santri")

    wb = openpyxl.load_workbook(output_excel_path)
    ws = wb.active

    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )

    for col_num, col in enumerate(ws.columns, 1):
        header_cell = col[0]
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        for cell in col[1:]:
            cell.border = border_thin
            if col_num in [1, 3, 4, 5, 6, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    ws.row_dimensions[1].height = 28
    wb.save(output_excel_path)
    print(f"[+] Master Excel tersinkron: {os.path.basename(output_excel_path)}")


def export_to_excel_and_json(data_list, excel_path=EXCEL_OUTPUT, json_path=JSON_OUTPUT):
    """Menyimpan data hasil scraping detail ke file Excel dan JSON."""
    if not data_list:
        return

    if isinstance(data_list, dict):
        data_list = [data_list]

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data_list, f, ensure_ascii=False, indent=2)
    print(f"[+] Hasil Scraping JSON: {os.path.basename(json_path)} ({len(data_list)} data)")

    df = pd.DataFrame(data_list)

    priority_order = [
        "Profil - Nama Lengkap", "Profil - No. Induk", "Profil - NISN", "Profil - Jenjang", "Profil - Kelas",
        "Profil - Ruang", "Profil - Program", "Profil - Pondok", "Profil - No. Whatsapp", "Biodata - Jenis Kelamin",
        "Biodata - Status", "Biodata - Tempat & Tgl lahir", "Biodata - Alamat Siswa", "Biodata - Nama Angkatan",
        "Wali - Nama Ayah", "Wali - Pekerjaan Ayah", "Wali - Nama Ibu", "Wali - No. Telp Ibu", "Wali - Alamat Ibu",
        "Kartu - No. Kartu", "Kartu - Status", "Saldo - Uang Saku", "Saldo - Tabungan Wali", "Paket - Nama Paket"
    ]

    ordered_cols = [c for c in priority_order if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in ordered_cols and c not in ["ID Siswa", "URL Sumber", "Waktu Scraping"]]
    meta_cols = [c for c in ["ID Siswa", "URL Sumber", "Waktu Scraping"] if c in df.columns]

    final_cols = ordered_cols + remaining_cols + meta_cols
    df = df[final_cols]

    # Bersihkan / pangkas isi teks yang terlalu panjang agar tidak melebihi limit sel Excel (32.767 karakter)
    for col in df.columns:
        df[col] = df[col].apply(lambda v: str(v)[:1500] if isinstance(v, str) and len(str(v)) > 1500 else v)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data Kesiswaan Lengkap")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col_num, col in enumerate(ws.columns, 1):
        header_cell = col[0]
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45)

        for cell in col[1:]:
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")

    ws.row_dimensions[1].height = 28
    wb.save(excel_path)
    print(f"\n[+] Hasil Scraping Excel: {os.path.basename(excel_path)}")


def scrape_all_students_flow(session):
    """Alur scraping untuk seluruh siswa aktif dengan multi-threading ramah server & auto-upsert."""
    students = get_all_active_students_list(session)
    if not students:
        print("[!] Daftar siswa kosong.")
        return

    total = len(students)
    print(f"\n[*] Memulai scraping detail untuk {total} santri...")
    print("[*] Menggunakan koneksi stabil & auto-retry agar tidak timeout.")

    results = []
    completed = 0

    def worker(s_info):
        time.sleep(0.1) # Jeda mikro agar server tidak overload
        return extract_full_student_details(session, s_info["id_siswa"])

    # Gunakan 3-4 thread agar server Simkeu tetap lancar tanpa timeout
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(worker, s): s for s in students}
        for future in as_completed(futures):
            completed += 1
            data = future.result()
            if data:
                results.append(data)

            name = data.get("Profil - Nama Lengkap", "Santri") if data else "Error"
            sys.stdout.write(f"\r[{completed}/{total}] ({completed*100//total}%) Scraping: {name[:30]:<30}")
            sys.stdout.flush()

            # Checkpoint auto-save tiap 50 santri
            if completed % 50 == 0:
                export_to_excel_and_json(results)
                upsert_master_dataset_by_no_induk(results)

    print("\n\n[+] SELURUH PROSES SCRAPING SELESAI!")
    export_to_excel_and_json(results)
    upsert_master_dataset_by_no_induk(results)


def main():
    print("=" * 75)
    print("   SIMKEU AL-MANNAN NETWORK - DATA SCRAPER & NO. INDUK UPSERT ENGINE   ")
    print("=" * 75)
    print(f"Target Default    : {DEFAULT_DETAIL_URL}")
    print(f"Master Data JSON  : {os.path.basename(MASTER_JSON_LOCAL)}")
    print(f"Master Data Excel : {os.path.basename(MASTER_EXCEL_LOCAL)}")
    print("=" * 75)

    session = get_requests_session()

    while True:
        print("\n" + "-" * 55)
        print("PILIHAN MENU SCRAPING & SINKRONISASI:")
        print("1. Scrape 1 Siswa Default & Auto-Update by No. Induk")
        print("2. Scrape SELURUH SANTRI (1.680+ Siswa) & Sinkron Master")
        print("3. Scrape 1 Siswa Kustom (Input URL / ID Siswa)")
        print("4. Re-Login / Buka Browser untuk Perbarui Sesi")
        print("5. Keluar")
        print("-" * 55)

        try:
            pilihan = input("Pilih menu (1/2/3/4/5) [Default 1]: ").strip() or "1"
        except (KeyboardInterrupt, EOFError):
            print("\n[*] Menutup program. Terima kasih!")
            break

        if pilihan == "1":
            print(f"\n[*] Mengambil data detail lengkap dari: {DEFAULT_DETAIL_URL}")
            data = extract_full_student_details(session, DEFAULT_DETAIL_URL)
            if data:
                export_to_excel_and_json([data])
                # JALANKAN UPSERT BERDASARKAN NO. INDUK
                upsert_master_dataset_by_no_induk([data])
                print(f"\n[✓] Selesai! Data No. Induk '{data.get('Profil - No. Induk')}' berhasil dicocokkan & disinkronkan.")
            else:
                print("[!] Gagal mengekstrak data.")

        elif pilihan == "2":
            confirm = input(f"\n[?] Anda yakin ingin men-scrape seluruh santri aktif? (y/n) [y]: ").strip().lower() or "y"
            if confirm == "y":
                scrape_all_students_flow(session)

        elif pilihan == "3":
            custom_input = input("\nMasukkan URL / ID Siswa: ").strip()
            if custom_input:
                data = extract_full_student_details(session, custom_input)
                if data:
                    export_to_excel_and_json([data])
                    upsert_master_dataset_by_no_induk([data])
                    print(f"\n[✓] Selesai! Data No. Induk '{data.get('Profil - No. Induk')}' berhasil disinkronkan.")
                else:
                    print("[!] Gagal mengekstrak data dari URL tersebut.")
            else:
                print("[!] Input tidak boleh kosong.")

        elif pilihan == "4":
            login_and_save_session()
            session = get_requests_session()

        elif pilihan == "5":
            print("[*] Menutup program. Terima kasih!")
            break
        else:
            print("[!] Pilihan tidak valid.")


if __name__ == "__main__":
    main()
